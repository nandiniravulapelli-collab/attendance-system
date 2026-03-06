from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework import status
from django.contrib.auth import login
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from openpyxl import load_workbook, Workbook
import datetime
import re
from decimal import Decimal
from .models import User, Attendance, Department, Subject, Section
from .serializers import (
    RegisterSerializer, LoginSerializer, AttendanceSerializer, UserSerializer,
    DepartmentSerializer, SubjectSerializer, SectionSerializer,
)


@api_view(['POST'])
@permission_classes([AllowAny])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


@csrf_exempt
@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):

    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        login(request, user)

        role = user.role
        if user.is_superuser and request.data.get("role") == "admin":
            role = "admin"

        return Response({
            "message": "Login successful",
            "id": user.id,
            "email": user.email,
            "role": role,
            "username": user.username,
            "full_name": user.full_name or user.username,
            "department": user.department or "",
        })

    return Response(serializer.errors, status=400)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def attendance_view(request):

    if request.method == 'GET':
        if request.user.role == 'student':
            records = Attendance.objects.filter(student=request.user)
        else:
            records = Attendance.objects.all()

        # Optional date range filtering: from_date, to_date in YYYY-MM-DD format
        from_date_str = request.query_params.get('from_date') or request.GET.get('from_date')
        to_date_str = request.query_params.get('to_date') or request.GET.get('to_date')
        from_date = to_date = None
        try:
            if from_date_str:
                from_date = datetime.date.fromisoformat(from_date_str)
            if to_date_str:
                to_date = datetime.date.fromisoformat(to_date_str)
        except (TypeError, ValueError):
            # Ignore invalid filters and return full data instead of erroring out
            from_date = to_date = None

        if from_date:
            records = records.filter(date__gte=from_date)
        if to_date:
            records = records.filter(date__lte=to_date)

        records_list = list(records.values('status', 'hours', 'total_hours'))
        total_attended = Decimal('0')
        total_scheduled = Decimal('0')
        for r in records_list:
            th = r.get('total_hours')
            h = r.get('hours')
            if th is not None and th > 0:
                total_scheduled += th
                total_attended += (h if h is not None else (Decimal('1') if str(r.get('status') or '').lower() == 'present' else Decimal('0')))
            else:
                total_scheduled += Decimal('1')
                total_attended += Decimal('1') if str(r.get('status') or '').lower() == 'present' else Decimal('0')

        # 1 hour = 1 class: total_classes and present_count are in "class" units (hours)
        total_classes = int(round(total_scheduled))
        present_count = int(round(total_attended))

        percentage = 0
        if total_scheduled > 0:
            percentage = float(total_attended / total_scheduled * 100)

        serializer = AttendanceSerializer(records, many=True)

        return Response({
            "total_classes": total_classes,
            "present_count": present_count,
            "total_attended_hours": float(total_attended),
            "total_hours": float(total_scheduled),
            "attendance_percentage": round(percentage, 2),
            "records": serializer.data
        })

    elif request.method == 'POST':
        if request.user.role not in ['faculty', 'admin']:
            return Response({"error": "Not authorized"}, status=403)

        data = request.data
        if isinstance(data, list):
            def parse_date_for_post(value):
                if value is None:
                    return None
                if isinstance(value, datetime.date):
                    return value if not isinstance(value, datetime.datetime) else value.date()
                if isinstance(value, datetime.datetime):
                    return value.date()
                try:
                    return datetime.date.fromisoformat(str(value).strip())
                except (TypeError, ValueError):
                    return None

            saved_count = 0
            errors = []
            for i, item in enumerate(data):
                if not isinstance(item, dict):
                    errors.append({"index": i, "errors": {"detail": "Invalid item, expected object."}})
                    continue
                try:
                    student_id = int(item.get('student'))
                except (TypeError, ValueError):
                    student_id = None
                subject = item.get('subject')
                date_val = item.get('date')
                status_val = item.get('status')
                if student_id is None or not str(subject).strip():
                    errors.append({"index": i, "errors": {"detail": "Missing student or subject."}})
                    continue
                if not str(status_val).strip().lower() in ('present', 'absent'):
                    errors.append({"index": i, "errors": {"detail": "Status must be 'present' or 'absent'."}})
                    continue
                date_obj = parse_date_for_post(date_val)
                if not date_obj:
                    errors.append({"index": i, "errors": {"detail": f"Invalid date: {date_val}"}})
                    continue
                if not User.objects.filter(pk=student_id, role='student').exists():
                    errors.append({"index": i, "errors": {"detail": f"Student id {student_id} not found."}})
                    continue
                status_normalized = str(status_val).strip().lower()
                if status_normalized in ('present', 'p'):
                    status_normalized = 'present'
                elif status_normalized in ('absent', 'a'):
                    status_normalized = 'absent'
                defaults = {"status": status_normalized}
                hours_val = item.get('hours')
                total_hours_val = item.get('total_hours')
                h_num, th_num = None, None
                if hours_val is not None and str(hours_val).strip() != '':
                    try:
                        h_num = float(str(hours_val).strip().replace(',', '.'))
                        if h_num >= 0:
                            defaults["hours"] = round(h_num, 2)
                    except (TypeError, ValueError):
                        pass
                if total_hours_val is not None and str(total_hours_val).strip() != '':
                    try:
                        th_num = float(str(total_hours_val).strip().replace(',', '.'))
                        if th_num > 0:
                            defaults["total_hours"] = round(th_num, 2)
                    except (TypeError, ValueError):
                        pass
                if defaults.get("hours") is not None and defaults.get("total_hours") is None:
                    defaults["total_hours"] = 1
                if defaults.get("total_hours") is not None and defaults.get("hours") is None:
                    defaults["hours"] = 1 if status_normalized == 'present' else 0
                _, created = Attendance.objects.update_or_create(
                    student_id=student_id,
                    subject=str(subject).strip(),
                    date=date_obj,
                    defaults=defaults,
                )
                saved_count += 1
            return Response({
                "created": saved_count,
                "records": [],  # frontend refetches via GET
                "errors": errors if errors else None,
            }, status=201 if saved_count > 0 else (400 if errors else 201))

        serializer = AttendanceSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=201)
        return Response(serializer.errors, status=400)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def user_list_view(request):
    role = request.query_params.get('role', 'student')
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    is_faculty = request.user.role == 'faculty'

    if not is_admin and not is_faculty:
        return Response({"detail": "Not allowed to list users."}, status=403)
    if is_faculty and role != 'student':
        return Response({"detail": "Faculty can only list students."}, status=403)

    qs = User.objects.filter(role=role).order_by('id')
    if is_faculty:
        if request.user.department:
            qs = qs.filter(department=request.user.department)
            section = request.query_params.get('section', '').strip()
            if section:
                qs = qs.filter(section=section)
            year = request.query_params.get('year', '').strip()
            if year:
                qs = qs.filter(year=year)
        else:
            qs = User.objects.none()
    elif is_admin and role == 'student':
        # Admin Mark Attendance: filter by department, section, year when provided
        department = request.query_params.get('department', '').strip()
        if department:
            qs = qs.filter(department=department)
        section = request.query_params.get('section', '').strip()
        if section:
            qs = qs.filter(section=section)
        year = request.query_params.get('year', '').strip()
        if year:
            qs = qs.filter(year=year)

    serializer = UserSerializer(qs, many=True)
    result = serializer.data
    if is_admin and role in ('student', 'faculty'):
        for i, u in enumerate(qs):
            if u.visible_password:
                result[i]['visible_password'] = u.visible_password
    return Response(result)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def user_detail_view(request, pk):
    try:
        target = User.objects.get(pk=pk)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=404)

    is_admin = request.user.role == 'admin' or request.user.is_superuser
    is_faculty = request.user.role == 'faculty'
    is_self = request.user.id == target.id
    faculty_can_edit_student = is_faculty and target.role == 'student' and target.department == request.user.department

    if not (is_admin or is_self or faculty_can_edit_student):
        return Response({"detail": "Not allowed to access this user."}, status=403)

    if request.method == 'GET':
        serializer = UserSerializer(target)
        data = serializer.data
        if is_admin and target.role in ('student', 'faculty') and target.visible_password:
            data['visible_password'] = target.visible_password
        return Response(data)

    elif request.method == 'PATCH':
        data = request.data.copy()
        new_password = (data.pop('new_password', None) or data.pop('password', None) or '').strip()
        current_password = (data.pop('current_password', None) or '').strip()

        can_change_password = False
        if new_password:
            if is_self and target.role in ('student', 'faculty'):
                can_change_password = True
                if not current_password:
                    return Response({"current_password": "Required when changing your own password."}, status=400)
                if not target.check_password(current_password):
                    return Response({"current_password": "Current password is incorrect."}, status=400)
            elif is_self and (target.role == 'admin' or target.is_superuser):
                can_change_password = True
                if not current_password:
                    return Response({"current_password": "Required when changing your own password."}, status=400)
                if not target.check_password(current_password):
                    return Response({"current_password": "Current password is incorrect."}, status=400)
            elif is_admin and target.role in ('student', 'faculty') and not target.is_superuser:
                can_change_password = True
            else:
                return Response({"detail": "You cannot change this user's password."}, status=403)

        if can_change_password and new_password:
            target.set_password(new_password)
            target.visible_password = new_password
            target.save(update_fields=['password', 'visible_password'])

        # Allow self to update username (must be unique)
        new_username = (data.pop('username', None) or '').strip()
        if is_self and new_username and new_username != target.username:
            if User.objects.filter(username=new_username).exclude(pk=target.pk).exists():
                return Response({"username": ["This username is already taken."]}, status=400)
            target.username = new_username
            target.save(update_fields=['username'])

        serializer = UserSerializer(target, data=data, partial=True)
        if serializer.is_valid():
            serializer.save()
            out = serializer.data
            if is_admin and target.role in ('student', 'faculty') and target.visible_password:
                out['visible_password'] = target.visible_password
            return Response(out)
        return Response(serializer.errors, status=400)

    elif request.method == 'DELETE':
        if not is_admin:
            return Response({"detail": "Only admin can delete users."}, status=403)
        if is_self:
            return Response({"detail": "You cannot delete your own account."}, status=400)
        if target.is_superuser:
            return Response({"detail": "Cannot delete superuser."}, status=400)
        target.delete()
        return Response(status=204)


# --- Departments (Branches) - Admin only ---
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def department_list_view(request):
    if request.user.role != 'admin' and not request.user.is_superuser:
        return Response({"detail": "Admin only."}, status=403)
    if request.method == 'GET':
        qs = Department.objects.all().order_by('code')
        return Response(DepartmentSerializer(qs, many=True).data)
    # POST
    serializer = DepartmentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def department_detail_view(request, pk):
    if request.user.role != 'admin' and not request.user.is_superuser:
        return Response({"detail": "Admin only."}, status=403)
    try:
        dept = Department.objects.get(pk=pk)
    except Department.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.method == 'GET':
        return Response(DepartmentSerializer(dept).data)
    if request.method == 'PATCH':
        serializer = DepartmentSerializer(dept, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    if request.method == 'DELETE':
        dept.delete()
        return Response(status=204)


# --- Sections: GET allowed for admin + faculty (for Mark Attendance dropdown); POST/DELETE admin only ---
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def section_list_view(request):
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    is_faculty = request.user.role == 'faculty'
    if request.method == 'POST' and not is_admin:
        return Response({"detail": "Admin only."}, status=403)
    if request.method == 'GET' and not is_admin and not is_faculty:
        return Response({"detail": "Not allowed."}, status=403)
    if request.method == 'GET':
        try:
            qs = Section.objects.all().order_by('name')
            return Response(SectionSerializer(qs, many=True).data)
        except Exception:
            return Response({"detail": "Sections table missing. Run: python manage.py migrate"}, status=500)
    # POST
    name = (request.data.get('name') or '').strip()
    if not name:
        return Response({"name": ["This field is required."]}, status=400)
    try:
        if Section.objects.filter(name=name).exists():
            return Response({"name": ["A section with this name already exists."]}, status=400)
        section = Section.objects.create(name=name)
        return Response(SectionSerializer(section).data, status=201)
    except Exception:
        return Response({"detail": "Could not save section. Run migrations: python manage.py migrate"}, status=500)


@api_view(['GET', 'DELETE'])
@permission_classes([IsAuthenticated])
def section_detail_view(request, pk):
    if request.user.role != 'admin' and not request.user.is_superuser:
        return Response({"detail": "Admin only."}, status=403)
    try:
        section = Section.objects.get(pk=pk)
    except Section.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.method == 'GET':
        return Response(SectionSerializer(section).data)
    if request.method == 'DELETE':
        section.delete()
        return Response(status=204)


# --- Subjects: GET allowed for admin + faculty (so faculty portal can load assigned subjects), POST admin only ---
@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def subject_list_view(request):
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    is_faculty = request.user.role == 'faculty'
    if request.method == 'POST' and not is_admin:
        return Response({"detail": "Admin only."}, status=403)
    if request.method == 'GET' and not is_admin and not is_faculty:
        return Response({"detail": "Not allowed."}, status=403)
    if request.method == 'GET':
        qs = Subject.objects.select_related('department').all().order_by('department__code', 'year', 'semester', 'code')
        department = request.query_params.get('department', '').strip()
        if department:
            qs = qs.filter(department__code=department)
        year = request.query_params.get('year', '').strip()
        if year:
            qs = qs.filter(year=year)
        semester = request.query_params.get('semester', '').strip()
        if semester:
            qs = qs.filter(semester=semester)
        return Response(SubjectSerializer(qs, many=True).data)
    # POST: require department (id or code), year and semester optional (default '1')
    data = request.data.copy()
    dept_id = data.get('department')
    if not dept_id:
        return Response({"department": ["This field is required."]}, status=400)
    dept = Department.objects.filter(pk=dept_id).first() or Department.objects.filter(code=dept_id).first()
    if not dept:
        return Response({"department": ["Department not found."]}, status=400)
    data['department'] = dept.id
    if data.get('year') in (None, ''):
        data['year'] = '1'
    if data.get('semester') in (None, ''):
        data['semester'] = '1'
    serializer = SubjectSerializer(data=data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(serializer.errors, status=400)


@api_view(['GET', 'PATCH', 'DELETE'])
@permission_classes([IsAuthenticated])
def subject_detail_view(request, pk):
    if request.user.role != 'admin' and not request.user.is_superuser:
        return Response({"detail": "Admin only."}, status=403)
    try:
        subj = Subject.objects.get(pk=pk)
    except Subject.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.method == 'GET':
        return Response(SubjectSerializer(subj).data)
    if request.method == 'PATCH':
        serializer = SubjectSerializer(subj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=400)
    if request.method == 'DELETE':
        subj.delete()
        return Response(status=204)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_student_upload_view(request):
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    if not is_admin:
        return Response({"detail": "Admin only."}, status=403)

    upload = request.FILES.get('file')
    if not upload:
        return Response({"detail": "No file uploaded. Use form field 'file'."}, status=400)
    if not str(upload.name).lower().endswith('.xlsx'):
        return Response({"detail": "Invalid file type. Please upload an .xlsx file."}, status=400)

    try:
        wb = load_workbook(filename=upload, data_only=True)
    except Exception:
        return Response({"detail": "Could not read Excel file. Make sure it is a valid .xlsx file."}, status=400)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Response({"detail": "Excel file is empty."}, status=400)

    header_row = rows[0]
    headers = [str(v).strip().lower() if v is not None else '' for v in header_row]
    required_cols = ['full_name', 'roll_number', 'email', 'department', 'section', 'year']
    missing = [c for c in required_cols if c not in headers]
    if missing:
        return Response(
            {
                "detail": "Missing required columns in header row.",
                "missing_columns": missing,
                "expected_columns": required_cols,
                "received_headers": headers,
            },
            status=400,
        )

    idx = {name: headers.index(name) for name in required_cols}

    created_count = 0
    skipped_existing = 0
    skipped_invalid = 0
    error_rows = []

    for row_number, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        if all((cell is None or str(cell).strip() == '') for cell in row):
            continue

        def _get(col_name):
            i = idx[col_name]
            if i >= len(row):
                return ''
            value = row[i]
            return '' if value is None else str(value).strip()

        full_name = _get('full_name')
        roll_number = _get('roll_number')
        email = _get('email')
        department = _get('department')
        section = _get('section')
        year = _get('year')

        if not roll_number or not email:
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": "Missing roll_number or email."})
            continue

        if User.objects.filter(roll_number=roll_number, role='student').exists():
            skipped_existing += 1
            continue

        username_base = roll_number or email.split('@')[0]
        username = username_base
        suffix = 1
        while User.objects.filter(username=username).exists():
            username = f"{username_base}_{suffix}"
            suffix += 1

        password = roll_number

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                role='student',
                full_name=full_name or username,
                roll_number=roll_number,
                department=department,
                section=section,
                year=str(year) if year is not None else '',
            )
            user.visible_password = password
            user.save(update_fields=['visible_password'])
        except Exception as exc:
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": f"Failed to create user: {exc.__class__.__name__}"})
            continue

        created_count += 1

    return Response(
        {
            "created": created_count,
            "skipped_existing": skipped_existing,
            "skipped_invalid": skipped_invalid,
            "total_processed_rows": len(rows) - 1,
            "errors": error_rows,
            "note": "New student accounts use roll_number as initial password.",
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_attendance_excel_view(request):
    is_admin = request.user.role == 'admin' or request.user.is_superuser
    if not is_admin:
        return Response({"detail": "Admin only."}, status=403)

    wb = Workbook()

    # Students sheet
    ws_students = wb.active
    ws_students.title = "Students"
    ws_students.append(
        ["Roll Number", "Name", "Email", "Department", "Section", "Year", "Phone"]
    )
    for u in User.objects.filter(role="student").order_by("id"):
        ws_students.append(
            [
                u.roll_number or "",
                u.full_name or u.username,
                u.email or "",
                u.department or "",
                u.section or "",
                u.year or "",
                u.phone or "",
            ]
        )

    # Faculty sheet
    ws_faculty = wb.create_sheet("Faculty")
    ws_faculty.append(["Name", "Email", "Department", "Phone"])
    for u in User.objects.filter(role="faculty").order_by("id"):
        ws_faculty.append(
            [
                u.full_name or u.username,
                u.email or "",
                u.department or "",
                u.phone or "",
            ]
        )

    # Admins sheet
    ws_admins = wb.create_sheet("Admins")
    ws_admins.append(["Name", "Email"])
    for u in User.objects.filter(role="admin").order_by("id"):
        ws_admins.append([u.full_name or u.username, u.email or ""])

    # Departments sheet
    ws_depts = wb.create_sheet("Departments")
    ws_depts.append(["Code", "Name"])
    for d in Department.objects.all().order_by("code"):
        ws_depts.append([d.code, d.name])

    # Subjects sheet
    ws_subjects = wb.create_sheet("Subjects")
    ws_subjects.append(["Code", "Name", "Department Code", "Year", "Semester"])
    for s in Subject.objects.select_related("department").all().order_by(
        "department__code", "year", "semester", "code"
    ):
        ws_subjects.append(
            [
                s.code,
                s.name,
                s.department.code if s.department_id else "",
                s.year,
                s.semester,
            ]
        )

    # Attendance sheet
    ws_att = wb.create_sheet("Attendance")
    ws_att.append(["Date", "Student Roll Number", "Subject", "Status", "attended_hours", "total_hours"])
    student_by_id = {
        u.id: u for u in User.objects.filter(role="student").only("id", "roll_number")
    }
    for a in Attendance.objects.all().order_by("date", "id"):
        student = student_by_id.get(a.student_id)
        ws_att.append(
            [
                a.date.isoformat() if a.date else "",
                student.roll_number if student and student.roll_number else "",
                a.subject or "",
                a.status or "",
                float(a.hours) if a.hours is not None else "",
                float(a.total_hours) if a.total_hours is not None else "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="attendance_data.xlsx"'
    wb.save(response)
    return response


def _parse_float_cell(value):
    """Parse a cell value to float; return None if empty or invalid."""
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return None
    try:
        s = str(value).strip().replace(',', '.')
        return round(float(s), 2)
    except (TypeError, ValueError):
        return None


def _parse_multi_float_cell(raw_value):
    """Parse a cell into a list of floats: single number -> [x], multiple (comma/semicolon separated) -> [x, y, ...]. Returns (list, None) or ([], error_msg)."""
    if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
        return [], "Missing value(s)."
    if isinstance(raw_value, (int, float)):
        try:
            v = round(float(raw_value), 2)
            return [v], None
        except (TypeError, ValueError):
            return [], f"Invalid number '{raw_value}'."
    text = str(raw_value).strip()
    parts = re.split(r'[,;\n]+', text)
    out = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        try:
            v = round(float(part.replace(',', '.')), 2)
            out.append(v)
        except (TypeError, ValueError):
            return [], f"Invalid number in '{part}'."
    if not out:
        return [], f"Invalid or missing number(s) in '{text}'."
    return out, None


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_attendance_upload_view(request):
    """Upload attendance in bulk from an Excel (.xlsx) file.

    Required columns: roll_number, subject, and either date (single) or dates (multiple).
    - date / dates: one or multiple dates (comma/semicolon separated). One record per date.
    - attended_hours / total_hours: one value (applies to all dates) or multiple (comma/semicolon separated)
      in the same order as dates, so each date gets the corresponding attended_hours and total_hours.
    Optional: status (when hours not provided). Duplicates are skipped. Returns created, skipped, errors.
    """
    if request.user.role not in ['admin', 'faculty'] and not request.user.is_superuser:
        return Response({"detail": "Only admin or faculty can upload attendance."}, status=403)

    upload = request.FILES.get('file')
    if not upload:
        return Response({"detail": "No file uploaded. Use form field 'file'."}, status=400)
    if not str(upload.name).lower().endswith('.xlsx'):
        return Response({"detail": "Invalid file type. Please upload an .xlsx file."}, status=400)

    try:
        wb = load_workbook(filename=upload, data_only=True)
    except Exception:
        return Response({"detail": "Could not read Excel file. Make sure it is a valid .xlsx file."}, status=400)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Response({"detail": "Excel file is empty."}, status=400)

    header_row = rows[0]
    headers = [str(v).strip().lower() if v is not None else '' for v in header_row]
    has_date_col = 'date' in headers
    has_dates_col = 'dates' in headers
    if not has_date_col and not has_dates_col:
        return Response(
            {
                "detail": "Missing required columns. Provide either 'date' (single date per row) or 'dates' (multiple dates in one cell, comma/semicolon separated).",
                "missing_columns": ["date or dates"],
                "expected_columns": ["roll_number", "subject", "date OR dates", "(status OR attended_hours + total_hours)"],
                "received_headers": headers,
            },
            status=400,
        )
    required_cols = ['roll_number', 'subject']
    missing = [c for c in required_cols if c not in headers]
    if missing:
        return Response(
            {
                "detail": "Missing required columns in header row.",
                "missing_columns": missing,
                "expected_columns": required_cols + ['date OR dates', '(status OR attended_hours + total_hours)'],
                "received_headers": headers,
            },
            status=400,
        )

    idx = {name: headers.index(name) for name in required_cols}
    idx['date'] = headers.index('date') if has_date_col else headers.index('dates')
    has_status = 'status' in headers
    has_attended_hours = 'attended_hours' in headers
    has_total_hours = 'total_hours' in headers
    if has_status:
        idx['status'] = headers.index('status')
    if has_attended_hours:
        idx['attended_hours'] = headers.index('attended_hours')
    if has_total_hours:
        idx['total_hours'] = headers.index('total_hours')

    # Backward compatibility: accept 'hours' as attended_hours if attended_hours not present
    if not has_attended_hours and 'hours' in headers:
        has_attended_hours = True
        idx['attended_hours'] = headers.index('hours')

    students_by_roll = {
        (u.roll_number or '').strip().upper(): u
        for u in User.objects.filter(role='student').exclude(roll_number__isnull=True)
    }

    created_count = 0
    skipped_existing = 0
    skipped_missing_student = 0
    skipped_missing_subject = 0
    skipped_invalid = 0
    error_rows = []

    def parse_date(value):
        if isinstance(value, datetime.date):
            return value if not isinstance(value, datetime.datetime) else value.date()
        if isinstance(value, datetime.datetime):
            return value.date()
        if value is None:
            return None
        if isinstance(value, (int, float)):
            try:
                serial = int(round(value))
                if serial < 1:
                    return None
                return (datetime.datetime(1899, 12, 31) + datetime.timedelta(days=serial)).date()
            except (ValueError, OverflowError):
                return None
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def parse_dates_cell(raw_value, row_number):
        """Return list of date objects from one cell: single date or multiple (comma/semicolon separated)."""
        if raw_value is None or (isinstance(raw_value, str) and not raw_value.strip()):
            return [], "Missing date(s)."
        if isinstance(raw_value, (int, float)):
            d = parse_date(raw_value)
            return ([d], None) if d else ([], f"Invalid date value '{raw_value}'.")
        if isinstance(raw_value, (datetime.date, datetime.datetime)):
            d = parse_date(raw_value)
            return ([d], None) if d else ([], f"Invalid date.")
        text = str(raw_value).strip()
        parts = re.split(r'[,;\n]+', text)
        out = []
        for part in parts:
            part = part.strip()
            if not part:
                continue
            d = parse_date(part)
            if not d and part.replace('.', '', 1).replace('-', '', 1).isdigit():
                try:
                    d = parse_date(float(part))
                except (TypeError, ValueError):
                    pass
            if d:
                out.append(d)
        if not out:
            return [], f"Invalid or missing date(s) in '{text}'."
        return out, None

    def _get(col_name):
        if col_name not in idx:
            return ''
        i = idx[col_name]
        if i >= len(row):
            return ''
        value = row[i]
        return '' if value is None else str(value).strip()

    for row_number, row in enumerate(rows[1:], start=2):
        if row is None:
            continue
        if all((cell is None or str(cell).strip() == '') for cell in row):
            continue

        roll_number = _get('roll_number')
        subject_raw = _get('subject')
        date_raw = row[idx['date']] if idx['date'] < len(row) else None

        if not roll_number or not subject_raw:
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": "Missing roll_number or subject."})
            continue

        student = students_by_roll.get(roll_number.strip().upper())
        if not student:
            skipped_missing_student += 1
            error_rows.append({"row": row_number, "reason": f"Student with roll_number '{roll_number}' not found."})
            continue

        subj = Subject.objects.filter(name__iexact=subject_raw).first() or Subject.objects.filter(code__iexact=subject_raw).first()
        if not subj:
            skipped_missing_subject += 1
            error_rows.append({"row": row_number, "reason": f"Subject '{subject_raw}' not found."})
            continue

        subject_value = subj.code or subj.name or subject_raw

        date_cell = row[idx['date']] if idx['date'] < len(row) else None
        date_list, date_err = parse_dates_cell(date_cell, row_number)
        if date_err or not date_list:
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": date_err or "Invalid or missing date(s)."})
            continue

        # Resolve attended_hours and total_hours: support single value or multiple (comma/semicolon separated) matching each date
        raw_attended = row[idx['attended_hours']] if has_attended_hours and idx['attended_hours'] < len(row) else None
        raw_total = row[idx['total_hours']] if has_total_hours and idx['total_hours'] < len(row) else None
        attended_list, attended_err = _parse_multi_float_cell(raw_attended) if (raw_attended is not None and str(raw_attended).strip() != '') else ([], None)
        total_list, total_err = _parse_multi_float_cell(raw_total) if (raw_total is not None and str(raw_total).strip() != '') else ([], None)

        if attended_err and has_attended_hours and (raw_attended is not None and str(raw_attended).strip() != ''):
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": attended_err})
            continue
        if total_err and has_total_hours and (raw_total is not None and str(raw_total).strip() != ''):
            skipped_invalid += 1
            error_rows.append({"row": row_number, "reason": total_err})
            continue

        # Build list of (date, attended, total) per record
        n_dates = len(date_list)
        use_hours = len(attended_list) > 0 and len(total_list) > 0

        if use_hours:
            if n_dates == len(attended_list) == len(total_list):
                date_hours_pairs = list(zip(date_list, attended_list, total_list))
            elif len(attended_list) == 1 and len(total_list) == 1:
                date_hours_pairs = [(d, attended_list[0], total_list[0]) for d in date_list]
            elif len(attended_list) != len(total_list):
                skipped_invalid += 1
                error_rows.append({"row": row_number, "reason": "attended_hours and total_hours must have the same number of values (or one value each for all dates)."})
                continue
            else:
                skipped_invalid += 1
                error_rows.append({"row": row_number, "reason": f"When using {n_dates} date(s), provide {n_dates} value(s) for attended_hours and total_hours, or one value each for all dates."})
                continue
        else:
            # Status-only path: default 1 hour per date
            if not has_status:
                skipped_invalid += 1
                error_rows.append({"row": row_number, "reason": "When attended_hours/total_hours are not both provided, status is required."})
                continue
            status_raw = _get('status')
            if not status_raw:
                skipped_invalid += 1
                error_rows.append({"row": row_number, "reason": "Status is required when attended_hours/total_hours are not both provided."})
                continue
            status_lower = status_raw.strip().lower()
            if status_lower in ['present', 'p']:
                status_value = 'present'
            elif status_lower in ['absent', 'a']:
                status_value = 'absent'
            else:
                skipped_invalid += 1
                error_rows.append({"row": row_number, "reason": f"Invalid status '{status_raw}'. Use Present/Absent."})
                continue
            attended_hours_value = 1 if status_value == 'present' else 0
            total_hours_value = 1
            date_hours_pairs = [(d, attended_hours_value, total_hours_value) for d in date_list]

        for date_value, attended_hours_value, total_hours_value in date_hours_pairs:
            if use_hours:
                if total_hours_value <= 0:
                    skipped_invalid += 1
                    error_rows.append({"row": row_number, "reason": "total_hours must be greater than 0."})
                    continue
                if attended_hours_value < 0 or attended_hours_value > total_hours_value:
                    skipped_invalid += 1
                    error_rows.append({"row": row_number, "reason": "attended_hours must be between 0 and total_hours."})
                    continue
                status_value = 'present' if attended_hours_value > 0 else 'absent'

            defaults = {
                "status": status_value,
                "hours": attended_hours_value,
                "total_hours": total_hours_value,
            }

            att_obj, created = Attendance.objects.get_or_create(
                student=student,
                subject=subject_value,
                date=date_value,
                defaults=defaults,
            )
            if created:
                created_count += 1
            else:
                skipped_existing += 1

    return Response(
        {
            "created": created_count,
            "skipped": skipped_existing + skipped_missing_student + skipped_missing_subject + skipped_invalid,
            "skipped_existing": skipped_existing,
            "skipped_missing_student": skipped_missing_student,
            "skipped_missing_subject": skipped_missing_subject,
            "skipped_invalid": skipped_invalid,
            "total_processed_rows": len(rows) - 1,
            "errors": error_rows,
        }
    )
